from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# กำหนด path สำหรับไฟล์ Excel 
EXCEL_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "playwright_pytest", "test_cases.xlsx")

# screenshots
SCREENSHOT_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "playwright_pytest", "screenshots")

os.makedirs(SCREENSHOT_DIR, exist_ok=True)      # if don't have screenshots => create


""" SAVE TO EXCEL """
def log_result(test_name, actual_result, page=None): # test case => login_pos01 , actual_result , page => screenshot
    print(f"🔹 Logging result for: {test_name}")

    # Have Excel ?
    if not os.path.exists(EXCEL_FILE):
        print("❌ ไม่พบไฟล์ test_cases.xlsx")
        return

    # load excel
    workbook = load_workbook(EXCEL_FILE)
    found = False   # fine test case ?

    # กำหนด target sheet
    if test_name.startswith("login"):
        target_sheet_name = "login01"
    elif test_name.startswith("logout"):
        target_sheet_name = "logout01"
    else:
        print(f"⚠️ ไม่สามารถระบุชีตสำหรับ {test_name}")
        workbook.close()
        return

    # ตรวจสอบว่า sheet ที่ต้องการมีอยู่ในไฟล์หรือไม่
    if target_sheet_name not in workbook.sheetnames:
        print(f"⚠️ ไม่พบชีต '{target_sheet_name}' ในไฟล์")
        workbook.close()
        return

    sheet = workbook[target_sheet_name]     # เลือก sheet ที่ต้องการใช้งาน

    # ตัด prefix ให้เหลือชื่อสั้น (เช่น pos01, neg02)
    short_test_name = test_name.split("_")[-1]

    # ค้นหาชื่อ test case ในคอลัมน์แรกของ sheet
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        test_case_cell = row[0] # A column
        
        ## เช็คว่า cell นั้นมีค่าและตรงกับ test_name แบบเต็ม หรือ ชื่อสั้น
        if test_case_cell.value and (test_case_cell.value.strip() == test_name or test_case_cell.value.strip() == short_test_name):
            row_number = test_case_cell.row                         # บันทึกแถวที่พบ
            status = "PASS" if "✅" in actual_result else "FAIL"    # status PASS have ✅ in actual_result ไม่เช่นนั้น FAIL

            # column save
            sheet[f"G{row_number}"] = actual_result
            sheet[f"H{row_number}"] = status

            # if have scr
            if page:
                screenshot_path = os.path.join(SCREENSHOT_DIR, f"{test_name}.png")   # กำหนด path สำหรับเก็บ screenshot ของ test case นี้
                page.screenshot(path=screenshot_path)                                # ถ่าย screenshot โดยใช้ method ของ page
                sheet[f"I{row_number}"] = screenshot_path                            # column save

                # have scr✅
                if os.path.exists(screenshot_path):
                    img = Image(screenshot_path)
                    img.width, img.height = 330, 210        # size of the screenshot
                    sheet.add_image(img, f"I{row_number}")  # column save

            found = True    #เจอ test case แล้ว
            print(f"✅ Updated '{test_name}' ในชีต '{target_sheet_name}' (พบจาก '{test_case_cell.value}')")
            break

    # ถ้าไม่พบ test case ที่ต้องการใน sheet
    if not found:
        print(f"⚠️ ไม่พบ Test Case ID '{test_name}' หรือ '{short_test_name}' ใน sheet '{target_sheet_name}'")

    # บันทึกไฟล์ Excel ที่แก้ไขแล้วและปิดไฟล์
    workbook.save(EXCEL_FILE)
    workbook.close()
    print(f"✅ Logging completed and saved in {EXCEL_FILE}")
